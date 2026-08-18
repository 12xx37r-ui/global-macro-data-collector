from __future__ import annotations

import json, math, os, statistics, io, csv, re, time
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import Any

import requests
from openpyxl import load_workbook

from global_m2 import build_global_m2

from treasury_card8 import (
    build_http_session, fetch_fred_all, finite, latest, values, change,
    annualized_index_change, clamp, mean, percentile, rmse, dm_test_squared_errors,
)

ROOT = Path(__file__).resolve().parent
OUT_DIR = ROOT / 'public' / 'data'

# Card 9: employment and consumption. Mostly official US + OECD global proxies.
CARD9_SERIES = {
    'UNRATE': {'label':'미국 실업률','core':True,'adverse_up':True},
    'PAYEMS': {'label':'미국 비농업고용','core':True,'adverse_up':False},
    'ICSA': {'label':'미국 신규실업수당','core':True,'adverse_up':True},
    'CCSA': {'label':'미국 계속실업수당','core':False,'adverse_up':True},
    'JTSJOL': {'label':'미국 구인건수','core':False,'adverse_up':False},
    'CES0500000003': {'label':'미국 시간당임금','core':False,'adverse_up':False},
    'RSAFS': {'label':'미국 소매판매','core':True,'adverse_up':False},
    'PCE': {'label':'미국 개인소비지출','core':True,'adverse_up':False},
    'UMCSENT': {'label':'미시간 소비심리','core':False,'adverse_up':False},
    'TOTALSL': {'label':'미국 소비자신용','core':False,'adverse_up':False},
    'PSAVERT': {'label':'미국 개인저축률','core':False,'adverse_up':False},
    'OECD_G20_CLI': {'label':'OECD G20 경기선행지수(CLI)','core':False,'adverse_up':False,'direct':True,'max_age_days':180},
}

# Card 10: commodities, energy, supply chain.
CARD10_SERIES = {
    'PALLFNFINDEXM': {'label':'IMF 글로벌 원자재 종합지수','core':True,'adverse_up':True},
    'PNRGINDEXM': {'label':'IMF 글로벌 에너지지수','core':True,'adverse_up':True},
    'DCOILWTICO': {'label':'WTI 유가','core':True,'adverse_up':True},
    'DCOILBRENTEU': {'label':'브렌트유','core':False,'adverse_up':True},
    'DHHNGSP': {'label':'미국 천연가스','core':False,'adverse_up':True},
    'PCOPPUSDM': {'label':'글로벌 구리가격','core':True,'adverse_up':False},
    'PALUMUSDM': {'label':'글로벌 알루미늄가격','core':False,'adverse_up':False},
    'PIORECRUSDM': {'label':'글로벌 철광석가격','core':False,'adverse_up':False},
    'PWHEAMTUSDM': {'label':'글로벌 밀가격','core':False,'adverse_up':True},
    'GSCPI': {'label':'뉴욕연은 글로벌 공급망압력','core':True,'adverse_up':True,'direct':True,'max_age_days':120},
    'PPIFIS': {'label':'미국 최종수요 생산자물가','core':True,'adverse_up':True},
}

# Card 12: futures market. Yahoo continuous contracts are delayed/free market inputs.
YAHOO_SYMBOLS = {
    'ES=F':'S&P500 선물','NQ=F':'나스닥100 선물','RTY=F':'러셀2000 선물',
    'ZT=F':'미국 2년 국채선물','ZF=F':'미국 5년 국채선물','ZN=F':'미국 10년 국채선물','ZB=F':'미국 30년 국채선물',
    'CL=F':'WTI 원유선물','GC=F':'금선물','HG=F':'구리선물','DX-Y.NYB':'미국 달러인덱스(현물 대체신호)','BTC=F':'비트코인선물',
}

HORIZONS = {'5d':5,'1m':21,'3m':63,'6m':126,'12m':252}


def now_iso(): return datetime.now(timezone.utc).isoformat()

GSCPI_XLSX_URL = "https://www.newyorkfed.org/medialibrary/research/interactives/gscpi/downloads/gscpi_data.xlsx"
GSCPI_CACHE_PATH = OUT_DIR / "cache" / "gscpi_official.json"
OECD_G20_CLI_URL = "https://sdmx.oecd.org/public/rest/data/OECD.SDD.STES,DSD_STES@DF_CLI,4.1/G20.M.LI...AA...H"
OECD_CLI_CACHE_PATH = OUT_DIR / "cache" / "oecd_g20_cli.json"


def _validate_xlsx_response(r: requests.Response) -> None:
    """Compatibility validator: accept both XLSX ZIP and legacy OLE/XLS."""
    body = r.content or b""
    content_type = (r.headers.get("content-type") or "").lower()
    if len(body) < 1024:
        raise ValueError(f"download too small ({len(body)} bytes)")
    if body[:2] == b"PK" or body[:8] == b"\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1":
        return
    preview = body[:80].decode("utf-8", errors="ignore").replace("\n", " ")
    raise ValueError(f"not an Excel workbook; content-type={content_type}; preview={preview!r}")


def _parse_dateish(value: Any, *, datemode: int | None = None) -> str | None:
    """Return ISO date for common monthly/date representations.

    The NY Fed GSCPI workbook has changed BIFF/XLS layouts over time.  Accept
    true Excel dates, numeric serials, YYYYMM/YYYMMDD numeric keys, and a broad
    set of textual date forms without adding another network source.
    """
    if value is None:
        return None
    if hasattr(value, "date"):
        try:
            return value.date().isoformat()
        except Exception:
            pass

    def _from_numeric(num: float) -> str | None:
        if not math.isfinite(num):
            return None
        # Numeric YYYYMM / YYYYMMDD keys are common in legacy statistical XLS.
        rounded = int(round(num))
        if abs(num - rounded) < 1e-8:
            text_num = str(abs(rounded))
            if 190001 <= rounded <= 210012 and len(text_num) == 6:
                y, m = divmod(rounded, 100)
                if 1 <= m <= 12:
                    return f"{y:04d}-{m:02d}-01"
            if 19000101 <= rounded <= 21001231 and len(text_num) == 8:
                y = rounded // 10000
                m = (rounded // 100) % 100
                d = rounded % 100
                try:
                    return datetime(y, m, d).date().isoformat()
                except ValueError:
                    pass
        # Excel serial date. Respect xlrd's workbook date system.
        if 1 <= num <= 80000:
            try:
                base = datetime(1904, 1, 1) if datemode == 1 else datetime(1899, 12, 30)
                dt = base + timedelta(days=num)
                if 1950 <= dt.year <= 2100:
                    return dt.date().isoformat()
            except Exception:
                pass
        return None

    if isinstance(value, (int, float)) and not isinstance(value, bool):
        parsed = _from_numeric(float(value))
        if parsed:
            return parsed

    text = str(value).strip().strip("'\"")
    if not text:
        return None
    text = (
        text.replace("\u2212", "-")
            .replace("\u2013", "-")
            .replace("\u2014", "-")
            .replace("\xa0", " ")
            .replace("\u200b", "")
            .strip()
    )

    # Numeric-as-text forms, including strings like '45123.0' or '202607.0'.
    if re.fullmatch(r"[+-]?\d+(?:\.0+)?", text):
        try:
            parsed = _from_numeric(float(text))
            if parsed:
                return parsed
        except ValueError:
            pass

    # Strip common time portions exported by Excel/BIFF.
    text = re.sub(r"[T ]\d{1,2}:\d{2}(?::\d{2}(?:\.\d+)?)?(?:Z|[+-]\d{2}:?\d{2})?$", "", text).strip()

    # Month keys such as 1997m1 / 1997M01 / 1997-01 / 199701.
    m = re.match(r"^(19\d{2}|20\d{2})\s*[mM]\s*(0?[1-9]|1[0-2])$", text)
    if m:
        return f"{int(m.group(1)):04d}-{int(m.group(2)):02d}-01"
    m = re.match(r"^(19\d{2}|20\d{2})[-_/. ](0?[1-9]|1[0-2])$", text)
    if m:
        return f"{int(m.group(1)):04d}-{int(m.group(2)):02d}-01"
    m = re.match(r"^(19\d{2}|20\d{2})(0[1-9]|1[0-2])$", text)
    if m:
        return f"{int(m.group(1)):04d}-{int(m.group(2)):02d}-01"
    m = re.match(r"^(19\d{2}|20\d{2})(0[1-9]|1[0-2])(0[1-9]|[12]\d|3[01])$", text)
    if m:
        try:
            return datetime(int(m.group(1)), int(m.group(2)), int(m.group(3))).date().isoformat()
        except ValueError:
            pass

    for fmt in (
        "%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d",
        "%m/%d/%Y", "%m/%d/%y", "%m-%d-%Y", "%m-%d-%y",
        "%d-%b-%Y", "%d-%b-%y", "%d-%B-%Y", "%d-%B-%y",
        "%d %b %Y", "%d %b %y", "%d %B %Y", "%d %B %y",
        "%Y-%m", "%Y/%m", "%Y.%m",
        "%b %Y", "%B %Y", "%b-%Y", "%B-%Y",
        "%b-%y", "%B-%y", "%b %y", "%B %y",
        "%b %d, %Y", "%B %d, %Y", "%d %b, %Y", "%d %B, %Y",
    ):
        try:
            dt = datetime.strptime(text, fmt)
            if 1950 <= dt.year <= 2100:
                if "%d" not in fmt:
                    dt = dt.replace(day=1)
                return dt.date().isoformat()
        except ValueError:
            continue
    return None


def _floatish(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    try:
        if isinstance(value, str):
            text = (
                value.strip()
                     .replace(",", "")
                     .replace("\u2212", "-")
                     .replace("\u2013", "-")
                     .replace("\u2014", "-")
                     .replace("\xa0", "")
                     .replace("\u200b", "")
            )
            if not text or text in {".", "..", "NA", "N/A", "nan", "null", "None"}:
                return None
            # Some legacy Excel exports leave a trailing footnote marker.
            text = re.sub(r"(?<=\d)[*†‡]+$", "", text)
            value = float(text)
        out = float(value)
    except (TypeError, ValueError):
        return None
    if not math.isfinite(out):
        return None
    return out


def _dedup_monthly_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    dedup: dict[str, dict[str, Any]] = {}
    for x in rows:
        date = str(x.get("date") or "")
        value = _floatish(x.get("value"))
        if len(date) < 7 or value is None:
            continue
        key = date[:7] + "-01"
        dedup[key] = {"date": key, "value": float(value)}
    return [dedup[k] for k in sorted(dedup)]


def _extract_gscpi_by_column_profile(matrix: list[list[Any]], *, datemode: int | None = None) -> list[dict[str, Any]]:
    """Infer date/value columns from row overlap instead of trusting sheet position.

    The live NY Fed legacy workbook currently exposes a four-column sheet with
    metadata rows before the observations.  This scorer tests every date/value
    column pair and keeps the pair with the most valid monthly observations.
    """
    if not matrix:
        return []
    width = max((len(r) for r in matrix), default=0)
    if width <= 1 or width > 64 or len(matrix) > 10000:
        return []

    best: list[dict[str, Any]] = []
    best_score = (-1, -1, -1)
    for dcol in range(width):
        for vcol in range(width):
            if dcol == vcol:
                continue
            candidate: list[dict[str, Any]] = []
            monotonic_pairs = 0
            prev_month = None
            for row in matrix:
                dv = row[dcol] if dcol < len(row) else None
                vv = row[vcol] if vcol < len(row) else None
                date_text = _parse_dateish(dv, datemode=datemode)
                val = _floatish(vv)
                if not date_text or val is None or not (-20.0 <= val <= 20.0):
                    continue
                month = date_text[:7]
                if prev_month is not None and month >= prev_month:
                    monotonic_pairs += 1
                prev_month = month
                candidate.append({"date": month + "-01", "value": val})
            candidate = _dedup_monthly_rows(candidate)
            # Prefer more rows, then chronological consistency, then adjacent columns.
            score = (len(candidate), monotonic_pairs, -abs(vcol - dcol))
            if score > best_score:
                best_score, best = score, candidate
    return best


def _extract_gscpi_rows_matrix_one_orientation(matrix: list[list[Any]], *, datemode: int | None = None) -> list[dict[str, Any]]:
    """Extract date/value pairs when observations run down rows."""
    if not matrix:
        return []

    header_row = None
    value_col = None
    date_col = None
    for ridx, row in enumerate(matrix[:80]):
        for cidx, cell in enumerate(row):
            label = str(cell or "").strip().lower()
            if "gscpi" in label or "global supply chain pressure" in label:
                header_row, value_col = ridx, cidx
                break
        if value_col is not None:
            for cidx, cell in enumerate(row):
                label = str(cell or "").strip().lower()
                if any(k in label for k in ("date", "month", "time", "period")):
                    date_col = cidx
                    break
            if date_col is None:
                date_col = max(0, value_col - 1)
            break

    rows: list[dict[str, Any]] = []
    if value_col is not None and header_row is not None:
        for row in matrix[header_row + 1:]:
            if value_col >= len(row):
                continue
            date_candidates = []
            if date_col is not None and date_col < len(row):
                date_candidates.append(row[date_col])
            date_candidates.extend(row[:min(len(row), max(8, value_col + 3))])
            date_text = next((p for p in (_parse_dateish(v, datemode=datemode) for v in date_candidates) if p), None)
            val = _floatish(row[value_col])
            if date_text and val is not None and -20.0 <= val <= 20.0:
                rows.append({"date": date_text, "value": val})

    # Generic row-wise fallback.
    if len(rows) < 36:
        fallback: list[dict[str, Any]] = []
        for row in matrix:
            date_candidates = [(idx, _parse_dateish(value, datemode=datemode)) for idx, value in enumerate(row)]
            date_candidates = [(idx, p) for idx, p in date_candidates if p]
            if not date_candidates:
                continue
            numeric = []
            for idx, value in enumerate(row):
                v = _floatish(value)
                if v is None or not (-20.0 <= v <= 20.0):
                    continue
                numeric.append((idx, v))
            for date_idx, date_text in date_candidates:
                values = [(idx, v) for idx, v in numeric if idx != date_idx]
                if values:
                    values.sort(key=lambda x: abs(x[0] - date_idx))
                    fallback.append({"date": date_text, "value": values[0][1]})
                    break
        if len(_dedup_monthly_rows(fallback)) > len(_dedup_monthly_rows(rows)):
            rows = fallback

    # Column-profile inference is specifically robust to metadata rows and
    # shifted Date/GSCPI columns in the live four-column legacy sheet.
    profiled = _extract_gscpi_by_column_profile(matrix, datemode=datemode)
    if len(profiled) > len(_dedup_monthly_rows(rows)):
        return profiled
    return _dedup_monthly_rows(rows)

def _transpose_matrix(matrix: list[list[Any]]) -> list[list[Any]]:
    if not matrix:
        return []
    width = max((len(r) for r in matrix), default=0)
    if width == 0:
        return []
    # Bound only absurdly wide sheets; the official series is ~350 monthly rows.
    if width > 5000 or len(matrix) > 5000:
        return []
    padded = [list(r) + [None] * (width - len(r)) for r in matrix]
    return [list(col) for col in zip(*padded)]


def _extract_gscpi_rows_matrix(matrix: list[list[Any]], *, datemode: int | None = None) -> list[dict[str, Any]]:
    """Extract GSCPI observations regardless of workbook orientation.

    NY Fed chart workbooks can be stored vertically (Date/GSCPI columns) or
    horizontally (dates across one row and values across another). Parse the
    original matrix first, then its transpose and keep the richer valid series.
    No network fallback is added here.
    """
    direct = _extract_gscpi_rows_matrix_one_orientation(matrix, datemode=datemode)
    transposed = _extract_gscpi_rows_matrix_one_orientation(_transpose_matrix(matrix), datemode=datemode)
    return transposed if len(transposed) > len(direct) else direct


def _sheet_shape_diagnostic(matrix: list[list[Any]]) -> str:
    rows = len(matrix)
    cols = max((len(r) for r in matrix), default=0)
    nonempty = sum(1 for r in matrix for v in r if v not in (None, ""))
    samples = []
    for r in matrix[:12]:
        vals = [str(v)[:40] for v in r[:12] if v not in (None, "")]
        if vals:
            samples.append("|".join(vals[:4]))
        if len(samples) >= 3:
            break
    return f"rows={rows},cols={cols},nonempty={nonempty},samples={samples}"


def _parse_gscpi_xlsx(content: bytes) -> list[dict[str, Any]]:
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    best: list[dict[str, Any]] = []
    for ws in wb.worksheets:
        matrix = [list(row) for row in ws.iter_rows(values_only=True)]
        parsed = _extract_gscpi_rows_matrix(matrix)
        if len(parsed) > len(best):
            best = parsed
    if len(best) < 36:
        raise ValueError(f"official XLSX contained only {len(best)} usable observations")
    return best


def _parse_gscpi_xls(content: bytes) -> list[dict[str, Any]]:
    import xlrd  # type: ignore
    book = xlrd.open_workbook(file_contents=content)
    best: list[dict[str, Any]] = []
    diagnostics: list[str] = []
    for sheet in book.sheets():
        matrix: list[list[Any]] = []
        for r in range(sheet.nrows):
            row: list[Any] = []
            for c in range(sheet.ncols):
                cell = sheet.cell(r, c)
                value = cell.value
                if cell.ctype == xlrd.XL_CELL_DATE:
                    try:
                        value = xlrd.xldate_as_datetime(value, book.datemode)
                    except Exception:
                        pass
                row.append(value)
            matrix.append(row)
        parsed = _extract_gscpi_rows_matrix(matrix, datemode=book.datemode)
        diagnostics.append(f"{sheet.name}:{_sheet_shape_diagnostic(matrix)}:parsed={len(parsed)}")
        if len(parsed) > len(best):
            best = parsed
    if len(best) < 36:
        diag = "; ".join(diagnostics[:4])
        raise ValueError(f"official XLS contained only {len(best)} usable observations; {diag}")
    return best


def _parse_gscpi_excel(content: bytes) -> list[dict[str, Any]]:
    if content[:2] == b"PK":
        return _parse_gscpi_xlsx(content)
    if content[:8] == b"\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1":
        return _parse_gscpi_xls(content)
    raise ValueError(f"unsupported official workbook signature={content[:8]!r}")

def _write_gscpi_cache(rows: list[dict[str, Any]]) -> None:
    GSCPI_CACHE_PATH.parent.mkdir(parents=True, exist_ok=True)
    GSCPI_CACHE_PATH.write_text(json.dumps({"saved_at_utc": now_iso(), "rows": rows}, ensure_ascii=False), encoding="utf-8")


def _read_gscpi_cache() -> list[dict[str, Any]]:
    try:
        payload = json.loads(GSCPI_CACHE_PATH.read_text(encoding="utf-8"))
        rows = payload.get("rows") or []
        return rows if len(rows) >= 36 else []
    except Exception:
        return []


def fetch_gscpi_official(session: requests.Session) -> tuple[list[dict[str, Any]], str]:
    """One official request, then local last-good only. No known-dead FRED retry."""
    try:
        r = session.get(
            GSCPI_XLSX_URL,
            timeout=(4, 12),
            allow_redirects=True,
            headers={
                "Accept": "application/vnd.ms-excel,application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,application/octet-stream;q=0.9,*/*;q=0.3",
                "Referer": "https://www.newyorkfed.org/research/policy/gscpi",
            },
        )
        r.raise_for_status()
        if len(r.content or b"") < 1024:
            raise ValueError(f"official workbook too small ({len(r.content or b'')} bytes)")
        rows = _parse_gscpi_excel(r.content)
        _write_gscpi_cache(rows)
        return rows, "New York Fed official workbook"
    except Exception as exc:
        cached = _read_gscpi_cache()
        if cached:
            return cached, f"local last-good cache (official refresh failed: {type(exc).__name__})"
        raise RuntimeError(f"official workbook/cache unavailable: {exc}")


def _read_oecd_cli_cache() -> list[dict[str, Any]]:
    try:
        payload = json.loads(OECD_CLI_CACHE_PATH.read_text(encoding="utf-8"))
        return list(payload.get("rows") or [])
    except Exception:
        return []


def _write_oecd_cli_cache(rows: list[dict[str, Any]]) -> None:
    OECD_CLI_CACHE_PATH.parent.mkdir(parents=True, exist_ok=True)
    OECD_CLI_CACHE_PATH.write_text(json.dumps({"saved_at_utc": now_iso(), "rows": rows}, ensure_ascii=False), encoding="utf-8")


def _oecd_start(previous_rows: list[dict[str, Any]]) -> tuple[str, str]:
    if not previous_rows:
        return "2010-01", "bootstrap"
    dates = [str(r.get("date") or "")[:7] for r in previous_rows if r.get("date")]
    dates = [d for d in dates if len(d) == 7]
    if not dates:
        return "2010-01", "bootstrap"
    latest = datetime.strptime(max(dates), "%Y-%m")
    # monthly release: 24-month overlap safely handles revisions with one request
    month_index = latest.year * 12 + latest.month - 1 - 24
    return f"{month_index//12:04d}-{month_index%12+1:02d}", "incremental"


def _parse_oecd_cli_csv(text: str) -> list[dict[str, Any]]:
    reader = csv.DictReader(io.StringIO(text))
    out: list[dict[str, Any]] = []
    for row in reader:
        period = row.get("TIME_PERIOD") or row.get("Time period") or row.get("TIME")
        raw = row.get("OBS_VALUE") or row.get("Observation value") or row.get("Value")
        if not period or raw in (None, "", ".."): continue
        try: value = float(raw)
        except (TypeError, ValueError): continue
        out.append({"date": str(period)[:7] + "-01", "value": value})
    dedup = {r["date"]: r for r in out}
    result = [dedup[d] for d in sorted(dedup)]
    if len(result) < 1:
        raise ValueError("OECD CLI CSV contained no usable observations")
    return result


def fetch_oecd_g20_cli(session: requests.Session) -> tuple[list[dict[str, Any]], str]:
    previous = _read_oecd_cli_cache()
    start_period, mode = _oecd_start(previous)
    try:
        r = session.get(
            OECD_G20_CLI_URL,
            params={"startPeriod": start_period, "dimensionAtObservation": "AllDimensions", "format": "csvfile"},
            timeout=(4, 12),
            headers={"Accept": "text/csv,*/*;q=0.2", "User-Agent": "global-macro-engine/2.6"},
        )
        r.raise_for_status()
        fresh = _parse_oecd_cli_csv(r.text)
        merged = _merge_direct_rows(previous, fresh)
        _write_oecd_cli_cache(merged)
        return merged, f"OECD Data Explorer SDMX ({mode}, one request)"
    except Exception as exc:
        if previous:
            return previous, f"local last-good cache (OECD refresh failed: {type(exc).__name__})"
        raise


def _merge_direct_rows(previous: list[dict[str, Any]], fresh: list[dict[str, Any]]) -> list[dict[str, Any]]:
    by = {str(r.get("date")): r for r in previous if r.get("date")}
    for row in fresh:
        if row.get("date"): by[str(row["date"])] = row
    return [by[d] for d in sorted(by)]


def fetch_card9_data(session: requests.Session) -> tuple[dict[str, list[dict[str, Any]]], list[str], dict[str, str]]:
    fred_ids = [sid for sid, meta in CARD9_SERIES.items() if not meta.get("direct")]
    data, errors = fetch_fred_all(session, fred_ids)
    source_overrides: dict[str, str] = {}
    try:
        data["OECD_G20_CLI"], source_overrides["OECD_G20_CLI"] = fetch_oecd_g20_cli(session)
    except Exception as exc:
        data["OECD_G20_CLI"] = []
        errors.append(f"OECD G20 CLI collection failed after official/cache attempt: {exc}")
    return data, errors, source_overrides


def fetch_card10_data(session: requests.Session) -> tuple[dict[str, list[dict[str, Any]]], list[str], dict[str, str]]:
    fred_ids = [sid for sid, meta in CARD10_SERIES.items() if not meta.get("direct")]
    data, errors = fetch_fred_all(session, fred_ids)
    source_overrides: dict[str, str] = {}
    try:
        data["GSCPI"], source_overrides["GSCPI"] = fetch_gscpi_official(session)
    except Exception as exc:
        data["GSCPI"] = []
        source_overrides["GSCPI"] = "New York Fed official workbook (unavailable)"
        errors.append(f"GSCPI collection failed after official workbook and local last-good cache attempts: {exc}")
    return data, errors, source_overrides

def aligned_group_index(histories: dict[str, list[dict[str, Any]]], symbols: list[str]) -> list[dict[str, Any]]:
    maps = {}
    for sym in symbols:
        rows = histories.get(sym, [])
        if len(rows) < 260:
            continue
        vals = [float(x["value"]) for x in rows]
        rets = {}
        for i in range(63, len(rows)):
            r21 = (vals[i] / vals[i-21] - 1) * 100 if vals[i-21] else 0.0
            r63 = (vals[i] / vals[i-63] - 1) * 100 if vals[i-63] else 0.0
            daily = [(vals[j] / vals[j-1] - 1) * 100 for j in range(max(1, i-62), i+1) if vals[j-1]]
            vol = statistics.stdev(daily) if len(daily) > 2 else 1.0
            rets[rows[i]["date"]] = clamp((0.6*r21 + 0.4*r63) / (vol*3 + 1), -3, 3)
        maps[sym] = rets
    dates = sorted(set().union(*[set(m) for m in maps.values()])) if maps else []
    out = []
    for d in dates:
        xs = [m[d] for m in maps.values() if d in m]
        if len(xs) >= max(1, len(maps)//2):
            out.append({"date": d, "value": 50 + 10*clamp(mean(xs), -3, 3)})
    return out

def build_futures_forecasts(histories: dict[str, list[dict[str, Any]]]) -> dict[str, Any]:
    groups = {
        "equity": ["ES=F", "NQ=F", "RTY=F"],
        "rates": ["ZT=F", "ZF=F", "ZN=F", "ZB=F"],
        "commodities": ["CL=F", "GC=F", "HG=F"],
        "crypto": ["BTC=F"],
    }
    group_forecasts = {}
    for group, syms in groups.items():
        idx = aligned_group_index(histories, syms)
        vals = values(idx)
        if len(vals) < 320:
            group_forecasts[group] = {"quality_gate": {"passed": False, "level": "참고용"}, "reason": "장기 검증표본 부족"}
            continue
        horizon_out = {}
        for key, steps, minimum in (("5d", 5, 180), ("1m", 21, 150), ("3m", 63, 100)):
            horizon_out[key] = walk_forward(vals, steps, minimum)
        group_forecasts[group] = {"current_index": vals[-1], "forecasts": horizon_out}
    passed = []
    for g, obj in group_forecasts.items():
        for h, fc in obj.get("forecasts", {}).items():
            if fc.get("quality_gate", {}).get("passed"):
                passed.append(f"{g}:{h}")
    return {"groups": group_forecasts, "passed_horizons": passed, "quality_gate": {"passed": bool(passed), "level": "기간별 검증 통과" if passed else "현재신호 중심"}}

def pctchg(vals:list[float], n:int)->float|None:
    if len(vals)<=n or vals[-1-n]==0: return None
    return (vals[-1]/vals[-1-n]-1)*100

def zscore(vals:list[float], window:int=120)->float:
    if len(vals)<10: return 0.0
    x=vals[-window:]
    m=statistics.mean(x); s=statistics.stdev(x) if len(x)>1 else 1.0
    return 0.0 if s==0 else (vals[-1]-m)/s

def monthly_last(rows:list[dict[str,Any]])->list[dict[str,Any]]:
    by={}
    for r in rows:
        if finite(r.get('value')) and r.get('date'): by[str(r['date'])[:7]]=r
    return [by[k] for k in sorted(by)]

def composite_history(data:dict[str,list[dict[str,Any]]], spec:dict[str,dict[str,Any]], mode:str)->list[dict[str,Any]]:
    # Monthly aligned rolling-z composite. Inputs are transformed to growth/momentum first.
    series_map={}
    for sid,meta in spec.items():
        rows=monthly_last(data.get(sid,[]))
        if len(rows)<36: continue
        vals=[float(x['value']) for x in rows]
        transformed=[]
        for i,r in enumerate(rows):
            if mode=='employment':
                if sid in ('UNRATE','ICSA','CCSA','PSAVERT'):
                    v=-(vals[i]-vals[max(0,i-3)])
                elif sid in ('PAYEMS','JTSJOL','RSAFS','PCE','TOTALSL','OECD_G20_CLI'):
                    v=(vals[i]/vals[max(0,i-3)]-1)*100 if vals[max(0,i-3)] else 0
                else:
                    v=(vals[i]/vals[max(0,i-3)]-1)*100 if vals[max(0,i-3)] else 0
            else:
                # Pressure: price/supply pressure up is adverse; copper/industrial metals retain growth signal.
                if sid in ('PCOPPUSDM','PALUMUSDM','PIORECRUSDM'):
                    v=(vals[i]/vals[max(0,i-3)]-1)*100 if vals[max(0,i-3)] else 0
                else:
                    v=-((vals[i]/vals[max(0,i-3)]-1)*100 if vals[max(0,i-3)] else 0)
            transformed.append(v)
        zs=[]
        for i,v in enumerate(transformed):
            w=transformed[max(0,i-60):i+1]
            m=statistics.mean(w); s=statistics.stdev(w) if len(w)>2 else 1
            zs.append(0 if s==0 else (v-m)/s)
        series_map[sid]={rows[i]['date'][:7]:zs[i] for i in range(len(rows))}
    dates=sorted(set().union(*[set(v) for v in series_map.values()])) if series_map else []
    out=[]
    for d in dates:
        vals=[]; weights=[]
        for sid,m in series_map.items():
            if d in m:
                vals.append(m[d]); weights.append(1.5 if spec[sid].get('core') else 1.0)
        if len(vals)>=3:
            score=sum(v*w for v,w in zip(vals,weights))/sum(weights)
            out.append({'date':d+'-01','value':50+clamp(score,-3,3)*10})
    return out

def candidates(train:list[float], h:int)->dict[str,float]:
    last=train[-1]
    def d(k): return (last-train[-1-k])/k if len(train)>k else 0
    m12=mean(train[-12:]); m24=mean(train[-24:]);
    return {
        'persistence':last,
        'short_trend':last+clamp(.65*d(1)+.35*d(3),-1.5,1.5)*h,
        'medium_trend':last+clamp(.25*d(1)+.45*d(3)+.30*d(6),-1.0,1.0)*h,
        'mean_reversion_12m':last+clamp((m12-last)*.18,-1.2,1.2)*min(h,6),
        'mean_reversion_24m':last+clamp((m24-last)*.12,-1.0,1.0)*min(h,8),
    }

def walk_forward(vals:list[float], h:int, min_samples:int=60)->dict[str,Any]:
    first=max(36,len(vals)-180-h)
    if len(vals)<first+h+15:
        return {'forecast':vals[-1],'model':'persistence','samples':0,'skill_pct':0,'direction_accuracy':None,'fallback_used':True,'range80':None}
    # Online/nested selection: the model used at each origin is selected only
    # from losses observed before that origin.  This prevents choosing the best
    # candidate on the same OOS sample later reported as its performance.
    candidate_losses:dict[str,list[float]]={}
    strategy_errors:list[float]=[]
    base_err:list[float]=[]
    strategy_hits:list[int]=[]
    selected_counts:dict[str,int]={}
    for o in range(first,len(vals)-h):
        tr=vals[:o+1]; actual=vals[o+h]; base=tr[-1]
        forecasts=candidates(tr,h)
        eligible={name:losses for name,losses in candidate_losses.items() if len(losses)>=12}
        selected=min(eligible,key=lambda k:mean(eligible[k])) if eligible else 'persistence'
        pred=forecasts[selected]
        selected_counts[selected]=selected_counts.get(selected,0)+1
        strategy_errors.append(pred-actual)
        base_err.append(base-actual)
        if abs(actual-base)>=1 and abs(pred-base)>=.5:
            strategy_hits.append(int((pred-base)*(actual-base)>0))
        for name,prediction in forecasts.items():
            candidate_losses.setdefault(name,[]).append((prediction-actual)**2)
    br=rmse(base_err); mr=rmse(strategy_errors); skill=(1-mr/br)*100 if br else 0
    fallback=skill<=0
    final_best=min(candidate_losses,key=lambda k:mean(candidate_losses[k])) if candidate_losses else 'persistence'
    if fallback: final_best='persistence'; mr=br; skill=0
    final=candidates(vals,h)[final_best]
    res=base_err if fallback else strategy_errors
    b80=percentile([abs(x) for x in res],.8)
    da=mean(strategy_hits) if strategy_hits else None
    dm=dm_test_squared_errors(res,base_err,max_lag=max(0,h-1))
    passed=len(res)>=min_samples and skill>=3.0 and da is not None and da>=.55 and dm.get('significant_5pct') and not fallback
    return {'forecast':final,'model':final_best,'samples':len(res),'rmse':mr,'baseline_rmse':br,'skill_pct':skill,
            'direction_accuracy':da,'fallback_used':fallback,'range80':[final-b80,final+b80] if finite(b80) else None,
            'selection_method':'online_prequential_candidate_selection','selected_model_counts':selected_counts,
            'quality_gate':{'passed':passed,'level':'독립검증 통과' if passed else '참고용/관망','dm_test':dm,
                            'requirements':{'skill_pct_min':3.0,'direction_accuracy_min':.55,'dm_p_value_max':.05,'hac_lag':max(0,h-1)}}}

def build_card9(session:requests.Session)->dict[str,Any]:
    data,errors,source_overrides=fetch_card9_data(session)
    hist=composite_history(data,CARD9_SERIES,'employment')
    vals=values(hist)
    if len(vals)<60: raise RuntimeError('Card9 composite history insufficient')
    forecasts={}
    for key,months in [('1m',1),('3m',3),('6m',6),('12m',12)]:
        forecasts[key]=walk_forward(vals,months,min_samples=60 if months<=6 else 45)
    current=vals[-1]
    f3=forecasts['3m']['forecast']; delta=f3-current
    signal='good' if delta>.8 else 'bad' if delta<-.8 else 'neutral'
    return {
        'schema_version':'1.1','engine_version':'card9-2.6.0-oecd-cli-direct','card':9,'title':'글로벌 고용·소비 경기','generated_at_utc':now_iso(),
        'current':current,'current_date':hist[-1]['date'],'forecast_3m':f3,'forecast_6m':forecasts['6m']['forecast'],
        'forecast_range80_3m':forecasts['3m']['range80'],'future_direction':'up' if delta>.4 else 'down' if delta<-.4 else 'flat',
        'market_signal':signal,'current_regime':'고용·소비 강함' if current>=55 else '고용·소비 약함' if current<45 else '고용·소비 보통',
        'future_regime':'회복' if signal=='good' else '둔화' if signal=='bad' else '보합',
        'investment_conclusion':'고용과 소비의 향후 방향을 경기민감주·내수·신용위험 판단에 반영합니다.',
        'forecasts':forecasts,'data_quality':quality(data,CARD9_SERIES,errors),
        'source_status':source_status(data,CARD9_SERIES,source_overrides),
        'model_specification':{'selection':'expanding walk-forward','benchmark':'persistence','inputs':list(CARD9_SERIES)},
    }

def build_card10(session:requests.Session)->dict[str,Any]:
    data,errors,source_overrides=fetch_card10_data(session)
    favorable_hist=composite_history(data,CARD10_SERIES,'commodity')
    # Public card semantics: higher value means greater commodity/energy/supply-chain pressure.
    hist=[{'date':x['date'],'value':100-float(x['value'])} for x in favorable_hist]
    vals=values(hist)
    if len(vals)<60: raise RuntimeError('Card10 composite history insufficient')
    forecasts={}
    for key,months in [('1m',1),('3m',3),('6m',6),('12m',12)]: forecasts[key]=walk_forward(vals,months,60 if months<=6 else 45)
    current=vals[-1]; f3=forecasts['3m']['forecast']; delta=f3-current
    # Pressure index semantics: a rise is adverse, a decline is favorable.
    signal='bad' if delta>.8 else 'good' if delta<-.8 else 'neutral'
    return {
        'schema_version':'1.1','engine_version':'card10-2.9.0-gscpi-column-profile-resilient','card':10,'title':'원자재·에너지·공급망 압력','generated_at_utc':now_iso(),
        'current':current,'current_date':hist[-1]['date'],'forecast_3m':f3,'forecast_6m':forecasts['6m']['forecast'],
        'forecast_range80_3m':forecasts['3m']['range80'],'future_direction':'up' if delta>.4 else 'down' if delta<-.4 else 'flat',
        'market_signal':signal,'current_regime':'공급·원가 부담' if current>=55 else '공급·원가 우호' if current<45 else '중립',
        'future_regime':'압력완화' if signal=='good' else '압력확대' if signal=='bad' else '보합',
        'investment_conclusion':'원가압력과 산업수요를 함께 봐 인플레이션·마진·원자재 자산 환경을 판단합니다.',
        'forecasts':forecasts,'data_quality':quality(data,CARD10_SERIES,errors),'source_status':source_status(data,CARD10_SERIES,source_overrides),
        'model_specification':{'selection':'expanding walk-forward','benchmark':'persistence','inputs':list(CARD10_SERIES)},
    }

def yahoo_history_with_snapshot(session:requests.Session,symbol:str)->tuple[list[dict[str,Any]],dict[str,Any]]:
    url='https://query1.finance.yahoo.com/v8/finance/chart/'+requests.utils.quote(symbol,safe='')
    # V218: force a network revalidation on every workflow run while retaining
    # 10y daily history for the existing forecasting/backtest model.
    params={'interval':'1d','range':'10y','events':'history','_ts':str(int(datetime.now(timezone.utc).timestamp()))}
    headers={'Cache-Control':'no-cache, no-store, max-age=0','Pragma':'no-cache'}
    r=session.get(url,params=params,headers=headers,timeout=(15,60)); r.raise_for_status()
    j=r.json(); z=j.get('chart',{}).get('result',[None])[0]
    if not z: return [],{}
    ts=z.get('timestamp') or []; close=((z.get('indicators') or {}).get('quote') or [{}])[0].get('close') or []
    out=[]
    for t,v in zip(ts,close):
        if finite(v): out.append({'date':datetime.fromtimestamp(t,timezone.utc).date().isoformat(),'value':float(v)})
    meta=z.get('meta') or {}
    market_time_utc=None
    try:
        raw_time=meta.get('regularMarketTime')
        if raw_time not in (None,''): market_time_utc=datetime.fromtimestamp(int(raw_time),timezone.utc).isoformat()
    except (TypeError,ValueError,OSError): market_time_utc=None
    price=meta.get('regularMarketPrice')
    try: price=float(price) if price is not None else None
    except (TypeError,ValueError): price=None
    snapshot={
        'symbol':symbol,'price':price,'market_time_utc':market_time_utc,
        'exchange':meta.get('exchangeName'),'exchange_timezone':meta.get('exchangeTimezoneName'),
        'market_state':meta.get('marketState'),'source':'Yahoo Finance chart metadata','retrieved_at_utc':now_iso(),'refetch_policy':'network_each_workflow_no_cache'
    }
    return out,snapshot

def yahoo_history(session:requests.Session,symbol:str)->list[dict[str,Any]]:
    # Backward-compatible helper retained for existing callers/tests.
    rows,_=yahoo_history_with_snapshot(session,symbol)
    return rows

def build_card12(session:requests.Session)->dict[str,Any]:
    histories={}; snapshots={}; errors=[]
    for idx,sym in enumerate(YAHOO_SYMBOLS):
        try:
            if idx:
                time.sleep(0.12)
            histories[sym],snapshots[sym]=yahoo_history_with_snapshot(session,sym)
        except Exception as e:
            histories[sym]=[]; snapshots[sym]={}; errors.append(f'{sym}: {e}')
    # Existing `current` schema is preserved, but its intended current-value
    # semantics now prefer the freshly fetched Yahoo metadata price. Daily history
    # remains unchanged and continues to drive all existing model calculations.
    current={}
    for sym,rows in histories.items():
        if not rows:
            continue
        base=latest(rows)
        snap=snapshots.get(sym) if isinstance(snapshots.get(sym),dict) else {}
        price=snap.get('price')
        try: price=float(price) if price is not None else None
        except (TypeError,ValueError): price=None
        market_date=base['date']
        try:
            if snap.get('market_time_utc'):
                market_date=datetime.fromisoformat(str(snap.get('market_time_utc')).replace('Z','+00:00')).date().isoformat()
        except Exception:
            pass
        current[sym]={'value':price if finite(price) else base['value'],'date':market_date}
    groups={'equity':['ES=F','NQ=F','RTY=F'],'rates':['ZT=F','ZF=F','ZN=F','ZB=F'],'commodities':['CL=F','GC=F','HG=F'],'dollar':['DX-Y.NYB'],'crypto':['BTC=F']}
    group_scores={}
    for g,syms in groups.items():
        comps=[]
        for symbol in syms:
            v=values(histories.get(symbol,[]))
            if len(v)>63:
                mom=(pctchg(v,21) or 0)*.6+(pctchg(v,63) or 0)*.4
                daily=[(v[i]/v[i-1]-1)*100 for i in range(max(1,len(v)-63),len(v)) if v[i-1]]
                vol=statistics.stdev(daily) if len(daily)>2 else 0
                comps.append(clamp(mom/(vol*3+1),-2,2))
        group_scores[g]=mean(comps) if comps else None
    composite=mean([x for x in [group_scores.get('equity'),group_scores.get('rates'),group_scores.get('commodities')] if finite(x)])
    signal='good' if composite>.25 else 'bad' if composite<-.25 else 'neutral'
    predictive=build_futures_forecasts(histories)
    return {'schema_version':'1.1','card':12,'title':'선물시장 현재신호·검증형 방향예측','generated_at_utc':now_iso(),
            'current':current,'group_scores':group_scores,'market_signal':signal,
            'current_regime':'위험선호 우세' if signal=='good' else '위험회피 우세' if signal=='bad' else '선물시장 혼조',
            'future_regime':'검증 통과 기간만 방향예측 사용',
            'market_implied_interpretation':'현재 선물가격·최근 추세·변동성이 암시하는 방향이며 미래 가격 자체가 아닙니다.',
            'predictive_validation':predictive,
            'investment_conclusion':'현재 선물시장 신호와 장기 순차 OOS를 통과한 기간별 방향예측을 분리해 사용합니다.',
            'data_quality':{'completeness':round(100*sum(bool(v) for v in histories.values())/len(histories),1),'warnings':errors},
            'source_status':{sym:{'ok':bool(histories[sym]),'label':YAHOO_SYMBOLS[sym],'latest':current.get(sym),'source':('Yahoo Finance delayed index proxy' if sym=='DX-Y.NYB' else 'Yahoo Finance delayed futures'),'market_snapshot':snapshots.get(sym) or {}} for sym in histories},
            # V217 additive overlay: same Yahoo request metadata, no new external call.
            'market_snapshots':snapshots,
            'freshness_contract':{'version':'V218','new_network_calls':0,'network_refetch_each_workflow':True,'http_cache_bypass':True,'existing_current_semantics_changed':False},
            'limitations':['무료 지연 선물자료와 달러인덱스 대체신호를 사용하며 거래소 실시간 유료 시세를 대체하지 않습니다.','선물가격은 미래 현물가격의 단순 예측값이 아닙니다.','검증 미통과 기간은 현재신호 또는 참고값으로만 사용합니다.']}

def _age_days(date_text: str | None) -> int | None:
    try:
        d = datetime.fromisoformat(str(date_text)[:10]).date()
        return (datetime.now(timezone.utc).date() - d).days
    except Exception:
        return None


def quality(data, spec, errors):
    core = [k for k, v in spec.items() if v.get('core')]
    stale = []
    fresh_ok = []
    core_fresh_ok = []
    for sid, meta in spec.items():
        row = latest(data.get(sid, []))
        age = _age_days((row or {}).get('date')) if row else None
        max_age = int(meta.get('max_age_days', 120))
        is_fresh = bool(row) and age is not None and age <= max_age
        fresh_ok.append(is_fresh)
        if meta.get('core'):
            core_fresh_ok.append(is_fresh)
        if row and not is_fresh:
            stale.append({'series': sid, 'latest_date': row.get('date'), 'age_days': age, 'max_age_days': max_age})
    warnings = list(errors)
    warnings.extend([f"{x['series']} 최신성이 기준 초과: {x['latest_date']} ({x['age_days']}일 경과)" for x in stale])
    return {
        'completeness': round(100 * sum(fresh_ok) / len(spec), 1),
        'raw_collection_completeness': round(100 * sum(bool(data.get(k)) for k in spec) / len(spec), 1),
        'core_completeness': round(100 * sum(core_fresh_ok) / len(core), 1),
        'missing_core': [k for k in core if not data.get(k)],
        'stale_series': stale,
        'warnings': warnings,
        'release_lag_policy': '월간·주간 자료는 공표시차와 계열별 최신성 기준을 함께 반영',
        'real_time_vintage': False,
    }


def source_status(data, spec, source_overrides=None):
    source_overrides = source_overrides or {}
    out = {}
    for sid, meta in spec.items():
        row = latest(data.get(sid, []))
        age = _age_days((row or {}).get('date')) if row else None
        max_age = int(meta.get('max_age_days', 120))
        out[sid] = {
            'ok': bool(data.get(sid)),
            'fresh': bool(row) and age is not None and age <= max_age,
            'age_days': age,
            'max_age_days': max_age,
            'label': meta['label'],
            'latest': row,
            'url': (GSCPI_XLSX_URL if sid == 'GSCPI' else OECD_G20_CLI_URL if sid == 'OECD_G20_CLI' else f'https://fred.stlouisfed.org/series/{sid}'),
            'collection_source': source_overrides.get(sid, 'FRED'),
        }
    return out

def load_json(path:Path)->dict[str,Any]:
    try:return json.loads(path.read_text(encoding='utf-8'))
    except:return {}

def _card8_validation(card8: dict[str, Any]) -> dict[str, Any]:
    gates = card8.get('quality_gates') or {}
    passed = []
    for horizon, gate in gates.items():
        targets = gate.get('passed_targets') or []
        if gate.get('passed') and targets:
            passed.append({'horizon': horizon, 'targets': targets})
    return {'passed': bool(passed), 'passed_horizons': passed, 'reason': '3·6개월 핵심금리 검증 통과 여부'}


def _generic_forecast_validation(card: dict[str, Any]) -> dict[str, Any]:
    passed = [h for h, fc in (card.get('forecasts') or {}).items() if (fc.get('quality_gate') or {}).get('passed')]
    return {'passed': bool(passed), 'passed_horizons': passed}


def _card12_validation(card12: dict[str, Any]) -> dict[str, Any]:
    pv = card12.get('predictive_validation') or {}
    passed = pv.get('passed_horizons') or []
    return {'passed': bool(passed), 'passed_horizons': passed}


def build_card11(card8, card9, card10, card12) -> dict[str, Any]:
    validations = {
        'card8': _card8_validation(card8),
        'card9': _generic_forecast_validation(card9),
        'card10': _generic_forecast_validation(card10),
        'card12': _card12_validation(card12),
    }
    signals = []
    weights = {'card8': 1.2, 'card9': 1.2, 'card10': 1.0, 'card12': 0.8}
    cards = {'card8': card8, 'card9': card9, 'card10': card10, 'card12': card12}
    for key, c in cards.items():
        s = c.get('market_signal')
        direction = 1 if s == 'good' else -1 if s == 'bad' else 0
        # A card contributes at full weight only when it has at least one validated horizon.
        confidence = 1.0 if validations[key]['passed'] else 0.5
        signals.append((direction, weights[key] * confidence))
    denom = sum(w for _, w in signals) or 1.0
    score = sum(s * w for s, w in signals) / denom * 100
    signal = 'good' if score >= 20 else 'bad' if score <= -20 else 'neutral'

    quality_checks = {
        'card8_validation': validations['card8']['passed'],
        'card9_validation': validations['card9']['passed'],
        'card10_validation': validations['card10']['passed'],
        'card12_validation': validations['card12']['passed'],
        'card8_data': card8.get('data_quality', {}).get('core_completeness', 0) >= 85,
        'card9_data': card9.get('data_quality', {}).get('core_completeness', 0) >= 85,
        'card10_data': card10.get('data_quality', {}).get('core_completeness', 0) >= 85,
        'card12_data': card12.get('data_quality', {}).get('completeness', 0) >= 80,
    }
    quality_passed = all(quality_checks.values())
    quality_level = '검증통과 신호 통합' if quality_passed else '조건부 통합판정'

    input_details = {
        'card8': {'label': '미국채 금리·실질금리', 'signal': card8.get('market_signal'), 'validation': validations['card8']},
        'card9': {'label': '글로벌 고용·소비', 'signal': card9.get('market_signal'), 'validation': validations['card9']},
        'card10': {'label': '원자재·에너지·공급망 압력', 'signal': card10.get('market_signal'), 'validation': validations['card10']},
        'card12': {'label': '선물시장 현재신호·방향예측', 'signal': card12.get('market_signal'), 'validation': validations['card12']},
    }
    return {
        'schema_version': '1.1', 'card': 11, 'title': '글로벌 경기국면 최종판정·투자환경', 'generated_at_utc': now_iso(),
        'score': round(score, 1), 'market_signal': signal,
        'current_regime': '확장·위험선호' if signal == 'good' else '수축·위험회피' if signal == 'bad' else '혼합·중립',
        'future_regime': '우호적 투자환경' if signal == 'good' else '불리한 투자환경' if signal == 'bad' else '중립적 투자환경',
        'asset_environment': {
            'global_equities': '우호' if signal == 'good' else '불리' if signal == 'bad' else '중립',
            'long_treasuries': '우호' if card8.get('market_signal') == 'good' else '중립',
            'gold': '우호' if card8.get('market_signal') == 'good' or card10.get('market_signal') == 'bad' else '중립',
            'commodities': '우호' if card10.get('market_signal') == 'good' and card9.get('market_signal') == 'good' else '중립',
            'cash_short_duration': '우호' if signal == 'bad' else '중립',
        },
        'inputs': {k: v.get('market_signal') for k, v in cards.items()},
        'input_details': input_details,
        'quality_gate': {'passed': quality_passed, 'level': quality_level, 'checks': quality_checks},
        'investment_conclusion': '검증을 통과한 하위 카드 신호와 최신성·완전성을 함께 반영해 자산군의 상대적 우호도를 판단합니다.',
    }

def main():
    session=build_http_session(); OUT_DIR.mkdir(parents=True,exist_ok=True)
    card9=build_card9(session); card10=build_card10(session); card12=build_card12(session)
    card8=load_json(OUT_DIR/'us_treasury_card8.json')
    card11=build_card11(card8,card9,card10,card12)
    global_m2 = build_global_m2(session)
    for n,obj in [(9,card9),(10,card10),(11,card11),(12,card12)]:
        (OUT_DIR/f'card{n}.json').write_text(json.dumps(obj,ensure_ascii=False,indent=2),encoding='utf-8')
    (OUT_DIR/'global_m2.json').write_text(json.dumps(global_m2,ensure_ascii=False,indent=2),encoding='utf-8')
    bundle={
        'generated_at_utc':now_iso(),
        'global_m2':global_m2,
        'macro':{'global_m2':global_m2},
        'cards':{'8':card8,'9':card9,'10':card10,'11':card11,'12':card12}
    }
    (OUT_DIR/'cards_8_12_bundle.json').write_text(json.dumps(bundle,ensure_ascii=False,indent=2),encoding='utf-8')

if __name__=='__main__': main()
